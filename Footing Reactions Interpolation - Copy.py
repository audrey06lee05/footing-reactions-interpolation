"""
=================Footing Reactions Interpolation========================
To use this script you will need to input the intial data from the
Footing Reaactions Interpolation Template - Copy.xlsx which is stored
in the same folder as this script. You may save the spreadsheet after
inputting the data and run this script. 
========================================================================
"""
import openpyxl as op
from openpyxl import load_workbook

# Load the spreadsheet
rd = op.load_workbook('Footing Reactions Interpolation Template - Copy.xlsx', data_only = True)     # rd = read data
sheet = rd.active
wb = op.load_workbook('Footing Reactions Interpolation Template - Copy.xlsx')
sheet1 = wb['Template']
ws = wb.active

#cell_value = sheet['H10'].value


def calculate_footings(max_WD, min_WD, max_leg_buoy, min_leg_buoy, no_WD_preload, no_WD_stillwater):
    WD_diff = float(max_WD) - float(min_WD)
    leg_buoy_diff = float(max_leg_buoy) - float(min_leg_buoy)
    pointOne_leg_buoy = 0

    if WD_diff != 0:
        pointOne_leg_buoy = float(leg_buoy_diff) / float(WD_diff * 10)

    WD = float(min_WD)
    count = 1

    while WD <= float(max_WD):
        WD += 0.1
        leg_buoy = float(min_leg_buoy) + (pointOne_leg_buoy * count)
        preload = float(no_WD_preload) - leg_buoy
        stillwater = float(no_WD_stillwater) - leg_buoy
        count += 1

        yield (WD, leg_buoy, preload, stillwater)


def main():
# Extracting data from the spreadsheet

    max_WD = sheet['H24'].value
    print("Maximum water depth: " + str(max_WD))
    min_WD = sheet['I24'].value
    print("Minimum water depth: " + str(min_WD) + '\n')

    max_leg_buoy_0m = sheet['H25'].value
    print('0m Penetration: \nLeg buoyancy (max WD): ' + str(max_leg_buoy_0m))
    min_leg_buoy_0m = sheet['I25'].value
    print("Leg buoyancy (min WD): " + str(min_leg_buoy_0m))

    max_preload_0m = sheet['H26'].value
    print("Preload footing reaction (max WD): " + str(max_preload_0m))
    min_preload_0m = sheet['I26'].value
    print("Preload footing reaction (min WD): " + str(min_preload_0m))

    max_stillwater_0m = sheet['H27'].value
    print("Stillwater footing reaction (max WD): " + str(max_stillwater_0m))
    min_stillwater_0m = sheet['I27'].value
    print("Stillwater footing reaction (min WD): " + str(min_stillwater_0m) + '\n')

    max_leg_buoy_30m = sheet['H11'].value
    print("30m Penetration: \nLeg buoyancy (max WD): " + str(max_leg_buoy_30m))
    min_leg_buoy_30m = sheet['I11'].value
    print("Leg buoyancy (min WD): " + str(min_leg_buoy_30m))

    max_preload_30m = sheet['H12'].value
    print("Preload footing reaction (max WD): " + str(max_preload_30m))
    min_preload_30m = sheet['I12'].value
    print("Preload footing reaction (min WD): " + str(min_preload_30m))

    max_stillwater_30m = sheet['H13'].value
    print("Stillwater footing reaction (max WD): " + str(max_stillwater_30m))
    min_stillwater_30m = sheet['I13'].value
    print("Stillwater footing reaction (min WD): " + str(min_stillwater_30m) + '\n' +
          "===========================================================================")

    no_WD_preload = min_preload_0m + min_leg_buoy_0m
    no_WD_stillwater = min_stillwater_0m + min_leg_buoy_0m

    check = input('Please check the data above, is it correct? Type Y to proceed: ')
    column_letter = 'M'


    if check.upper() == 'Y':
        max_row_for_column = max((cell.row for cell in ws[f'{column_letter}'] if cell.value is not None))

        for column_data in [(max_leg_buoy_0m, min_leg_buoy_0m, '0m penetration'),
                            (max_leg_buoy_30m, min_leg_buoy_30m, '30m penetration')]:
            max_leg_buoy, min_leg_buoy, penetration_label = column_data
            for data in calculate_footings(max_WD, min_WD, max_leg_buoy, min_leg_buoy, no_WD_preload, no_WD_stillwater):
                formatted_data = float(f'{data[0]:.1f}')
                WD_list = [formatted_data]
                for row in range(3, max_row_for_column + 1):
                    cell = ws[f'{column_letter}{row}'].value
                    if float(cell) in WD_list:
                        print(f"\n{penetration_label}: \nWater Depth: {cell:.1f}\nLeg Buoyancy: {data[1]:.1f}\nPreload Footing Reaction: {data[2]:.1f}\nStillwater Footing Reaction: {data[3]:.1f}\n")
                    #print("===========================================================================")


if __name__ == "__main__":
    main()
